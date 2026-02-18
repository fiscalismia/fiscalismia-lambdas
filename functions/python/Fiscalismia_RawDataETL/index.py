import json
import boto3
import tempfile
import requests # not in aws runtime
import time
from datetime import datetime
import zoneinfo
from io import BytesIO
from openpyxl import load_workbook
from openpyxl import Workbook
from aws_lambda_powertools import Logger
from aws_lambda_powertools.utilities import parameters
s3_client = boto3.client('s3')
s3_bucket = 'fiscalismia-raw-data-etl-storage'
logger = Logger(service="Fiscalismia_RawDataETL")
timedelta_analysis = []

def authenticate_request(body, headers):
  contentLength = int(headers.get('Content-Length', 0))
  authorization = headers.get('authorization', None)
  requestIp = headers.get('X-Forwarded-For', None)

  logger.info("Request received", extra={"ip": requestIp})

  secret = parameters.get_secret(
    "arn:aws:secretsmanager:eu-central-1:010928217051:secret:/api/fiscalismia/API_GW_SECRET_KEY-4AjIFN",
    transform="json"
  )
  secret_api_key = secret.get("API_GW_SECRET_KEY", None)
  # block access if payload is sent
  if body or contentLength > 0:
    return {
      "statusCode": 422,
      "body": json.dumps({"message": "No payload expected. Request body should be empty."})
    }
  # block access if authorization header does not include the API TOKEN
  if authorization == None or secret_api_key == None or authorization != secret_api_key:
    return {
      "statusCode": 403,
      "body": json.dumps({"message": "Invalid Authorization header."})
    }
  # Everything is fine. Proceed
  logger.info("Request authenticated successfully")
  return {"statusCode": 200}

def log_debug_info(event, headers, context, info_log=True):
  extractedEventKeys = {
    "HTTP" : event.get('httpMethod', None),
    "Host" : headers.get("Host", None),
    "path" : event.get('path', None),
    "contentLength" : int(headers.get('Content-Length', 0)),
    "X-Forwarded-For" : headers.get('X-Forwarded-For', None),
    "User-Agent" : headers.get('User-Agent', None),
    "queryStringParameters" : event.get('queryStringParameters', None),
    "body" : event.get("body", None)
  }
  if info_log:
    logger.info("Debug info extracted", extra={"eventKeys": extractedEventKeys})
  logger.debug("Debug info extracted", extra={"eventKeys": extractedEventKeys})

def add_time_analysis_entry(start_time, log_msg):
  timedelta_analysis.append(f"{round((time.time_ns() - start_time) / 1_000_000)}ms time passed after [{log_msg}]")

def log_time_analysis(info_log=True):
  if info_log:
    logger.info("Timedelta analysis concluded.", extra={ "timedelta_analysis" : timedelta_analysis })
  logger.debug("Timedelta analysis concluded.", extra={ "timedelta_analysis" : timedelta_analysis })

def lambda_handler(event, context):
  start_time = time.time_ns()
  add_time_analysis_entry(start_time, "function invocation")
  body = event.get("body", None)
  headers = event.get("headers", None)
  log_debug_info(event, headers, context)
  auth_response = authenticate_request(body, headers)
  if auth_response.get("statusCode", None) != 200:
    return auth_response
  add_time_analysis_entry(start_time, "request authentication")
  try:
    secret = parameters.get_secret(
      "arn:aws:secretsmanager:eu-central-1:010928217051:secret:/google/sheets/fiscalismia-datasource-url-k38OGm",
      transform="json"
    )
    # Verify spreadsheet url is not malformed
    sheet_url = secret["GOOGLE_SHEETS_URL"]
    if not sheet_url or "docs.google.com/spreadsheets" not in sheet_url:
      return {
        "statusCode": 400,
        "body": json.dumps({"error": "Missing spreadsheets url in lambda environment."})
      }
    if "pub?output=xlsx" in sheet_url or "export?format=xlsx" in sheet_url:
      pass
    elif "/edit" in sheet_url:
      sheet_url = sheet_url.split("/edit")[0] + "/export?format=xlsx"
    elif "/view" in sheet_url:
      sheet_url = sheet_url.split("/view")[0] + "/export?format=xlsx"
    elif "/pubhtml" in sheet_url:
      sheet_url = sheet_url.split("/pubhtml")[0] + "/pub?output=xlsx"

    # Download the spreadsheet from google docs into memory
    response = requests.get(sheet_url, stream=True, timeout=(3, 10)) # (3s connect timeout, 10s read timeout)
    add_time_analysis_entry(start_time, "request spreadsheet via URL")
    if response.status_code != 200:
      return {
        "statusCode": 500,
        "body": json.dumps({"error": "Failed to download the sheet"})
      }
    s3_buffer = BytesIO(response.content)
    workbook_buffer = BytesIO(response.content)

    # Load temporary file as binary object into memory for S3 backup
    berlin_tz = zoneinfo.ZoneInfo("Europe/Berlin")
    timestamp = datetime.now(tz=berlin_tz).strftime("%Y-%m-%d-%H-%M-%S")
    s3_key = f"tmp/{timestamp}-Fiscalismia-Datasource.xlsx"
    add_time_analysis_entry(start_time, "load sheet into memory temp file")
    s3_client.upload_fileobj(s3_buffer, s3_bucket, s3_key)
    logger.info(f"Worksheet persisted to s3://{s3_bucket}/{s3_key}")
    add_time_analysis_entry(start_time, "persist temp file to s3")

    # Load openpyxl workbook in memory for data extraction
    workbook = load_workbook(workbook_buffer, read_only=True, data_only=True, keep_links=False)
    add_time_analysis_entry(start_time, "loaded workbook into memory")
    logger.info("Loaded openpyxl workbook into memory")
    sheet_names = workbook.sheetnames

    # log timedeltas for performance monitoring
    logger.info("finalized extract transform loading operation")
    log_time_analysis()
    workbook.close()
    return {
      "statusCode": 200,
      "body": json.dumps({"sheets": sheet_names})
    }

  except Exception as e:
    return {
        "statusCode": 500,
        "body": json.dumps({"error": str(e)})
    }
